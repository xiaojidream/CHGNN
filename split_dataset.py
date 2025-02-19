import os
import sys
import pandas as pd
from openpyxl import Workbook
import random
import SimpleITK as sitk
import numpy as np
import cv2



def generate_data(data_path, data_list, output_name):
    pos_path = os.path.join(data_path, 'positive')
    neg_path = os.path.join(data_path, 'negative')

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'id'
    ws['B1'] = 'label'
    ws['C1'] = 'path'
    idx_row = 2

    for id in data_list:
        if os.path.exists(os.path.join(pos_path, id)):
            label = 1
            path = os.path.join(pos_path, id, 'A')
        else:
            label = 0
            path = os.path.join(neg_path, id, 'A')

        ws['A%d'%(idx_row)] = id
        ws['B%d' % (idx_row)] = label
        ws['C%d' % (idx_row)] = path
        idx_row = idx_row + 1
        wb.save(output_name)



def generate_z_data(data_path, data_list, output_name):
    pos_path = os.path.join(data_path, 'positive')
    neg_path = os.path.join(data_path, 'negative')

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'id'
    ws['B1'] = 'label'
    idx_row = 2

    for id in data_list:
        if os.path.exists(os.path.join(pos_path, id)):
            label = 1
        else:
            label = 0

        ws['A%d' % (idx_row)] = id
        ws['B%d' % (idx_row)] = label


        idx_row = idx_row + 1
        wb.save(output_name)



def main():
    save_path = r'.\data'
    internal_path = r'E:\yyx\RanK\OriginData\data\segdata'

    all_files = os.listdir(os.path.join(internal_path, 'positive')) + os.listdir(os.path.join(internal_path, 'negative'))
    train_list = []
    val_list = []

    train_ids = np.array(pd.read_excel(r'E:\yyx\RanK\bootstrap\random\train.xlsx')['ID'])
    unique_arr, index = np.unique(train_ids, return_index=True)
    # 根据索引排序得到去重后的数组
    train_ids = unique_arr[np.argsort(index)]

    val_ids = np.array(pd.read_excel(r'E:\yyx\RanK\bootstrap\random\val.xlsx')['ID'])
    unique_arr, index = np.unique(val_ids, return_index=True)
    # 根据索引排序得到去重后的数组
    val_ids = unique_arr[np.argsort(index)]


    for train_id in train_ids:
        for id in all_files:
            if id.split('_')[0] == train_id:
                train_list.append(id)
    for val_id in val_ids:
        for id in all_files:
            if id.split('_')[0] == val_id:
                val_list.append(id)

    generate_z_data(internal_path, train_list, os.path.join(save_path, 'train.xlsx'))
    generate_z_data(internal_path, val_list, os.path.join(save_path, 'val.xlsx'))

    gen_encoder_data(train_list, 'train')
    gen_encoder_data(val_list, 'val')


def gen_encoder_data(datas, name):

    save_path = r'E:\yyx\RanK\graph-slide-copy\data\encoder_data'

    npy_ids = []
    npy_labels = []
    npy_paths = []

    for i in range(len(datas)):
        id = datas[i]
        if os.path.exists(os.path.join(r'E:\yyx\RanK\OriginData\data\segdata\positive', id)):
            label = 'positive'
        else:
            label = 'negative'

        base_path = r'E:\yyx\RanK\OriginData\data\segdata'
        base_path = os.path.join(base_path, label, id, 'img_npy')
        npy_files = os.listdir(base_path)
        for file in npy_files:
            npy_label = file.split('-')[1]
            npy_ids.append(id)
            npy_labels.append(npy_label)
            npy_paths.append(os.path.join(base_path, file))



    data = {
        'id': npy_ids,
        'label': npy_labels,
        'path': npy_paths
    }
    npy_df = pd.DataFrame(data)
    npy_df.to_excel(os.path.join(save_path, name + '.xlsx'), index=False)








if __name__ == '__main__':
    main()


